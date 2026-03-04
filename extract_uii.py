#!/usr/bin/env python3
"""
extract_uii.py — Извлечение данных из списка УИИ (docx → xlsx)

Структуры таблиц:
  A  (9 кол): дата|№лд|ФИО+адрес|ДР|суд|обязанности|окончание|примечание
  B8 (8 кол): дата|№лд|ФИО+адрес|ДР|суд|обязанности|окончание+примечание
  OR (8 кол): дата|№лд|ФИО+адрес|ДР|суд — нет обязанностей и окончания
  IR (7 кол): дата|№лд|ФИО+адрес|ДР|суд|примечание — ИР, нет обязанностей/окончания
  OS (9 кол): дата|№лд|ФИО|ДР|МЖ_отдельно|суд|окончание — Отсрочка
  DA (8 кол): дата|№лд|ФИО+адрес|ДР|суд|обязанности — нет окончания (Домашний арест)
  ZODA(8кол): дата|№лд|ФИО+адрес|ДР|суд+обязанности_вместе — нет окончания
"""

import sys
import re
import os
from docx import Document
from docx.oxml.ns import qn
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Настройки ──────────────────────────────────────────────────────────────────
INPUT_FILE  = 'СПИСОК 15 ОП СЕНТЯБРЬ 2025.docx'
OUTPUT_FILE = 'Список_extracted.xlsx'

# ── Разбивка ФИО + адрес ───────────────────────────────────────────────────────

# ── Предобработка: разбивка склеенных слов ────────────────────────────────────
# 1) «Векилоглы[нет пробела]Гражданский» → строчная кириллица + заглавная = пробел
_MERGE_CYR = re.compile(r'([а-яё])([А-ЯЁ])')   # БЕЗ IGNORECASE (важно!)
# 2) «Александровичул.» → строчная буква + аббревиатура адреса без пробела
_MERGE_ABBR = re.compile(r'([а-яё])(ул\.|пр\.|д\.|кв\.|корп\.)')   # тоже без IGNORECASE

# Паттерн: начало строки = адрес
_ADDR_LINE_START = re.compile(
    r'^(?:'
    r'г\.?\s|г\s|'            # «г. СПб», «г Мурино»
    r'спб\b|'                 # СПб / Спб / спб
    r'санкт-петербург|'
    r'ло\b|'                  # Ленобласть
    r'зарег|прож|рф\b|'
    r'ул\.|пр\.|'             # без требования пробела
    r'пр-т|пр-п|'
    r'краснодар|мурман|'
    r'\d{1,3}-\d'             # «108-196» — формат строение-квартира
    r')',
    re.IGNORECASE,
)

# Паттерн: разделитель «ФИО, адрес» внутри одной строки
_ADDR_INLINE = re.compile(
    r'[,\.]\s*(?='
    r'(?:г\.?\s|г\s|спб\b|санкт-петербург|ло\b|'
    r'зарег|прож|рф\b|'
    r'ул\.|пр\.|пр-т|пр-п|'
    r'д\.|кв\.|корп\.|'       # «, д. 21», «, кв. 23»
    r'\d{1,3}-\d)'
    r')',
    re.IGNORECASE,
)


def _is_addr_line(line: str, fio_word_count: int) -> bool:
    """Возвращает True, если строка является адресом.

    fio_word_count — сколько слов ФИО уже собрано до этой строки.
    Проверки на кв./д. включаются только когда ФИО уже полное (≥ 3 слов),
    чтобы не поглощать патроним «Николаевич, ул. Верности...» целиком.
    """
    if _ADDR_LINE_START.match(line):
        return True
    if fio_word_count > 0:
        # «Луначарского 106-174» — строение-квартира без явного ключевого слова
        if re.search(r'\d{1,3}-\d{1,3}', line):
            return True
    if fio_word_count >= 3:
        # ФИО уже полное — строка с «д.» / «кв.» / «корп.» = адрес целиком
        if re.search(r'[,\s](?:д|кв|корп|к)\.', line, re.IGNORECASE):
            return True
    return False


def split_fio_address(text: str):
    """Разбивает ячейку «ФИО + адрес» на (ФИО, адрес)."""
    if not text:
        return '', ''

    # Предобработка: разбиваем склеенные слова
    text = _MERGE_CYR.sub(r'\1 \2', text)    # «ВекилоглыГражданский» → «Векилоглы Гражданский»
    text = _MERGE_ABBR.sub(r'\1 \2', text)   # «Александровичул.» → «Александрович ул.»

    lines = [l.strip() for l in text.split('\n') if l.strip()]
    fio_parts, addr_parts = [], []
    addr_started = False

    for line in lines:
        if addr_started:
            addr_parts.append(line)
            continue

        fio_word_count = len(' '.join(fio_parts).split())
        if _is_addr_line(line, fio_word_count):
            addr_started = True
            addr_parts.append(line)
            continue

        # Ищем разделитель внутри строки
        m = _ADDR_INLINE.search(line)
        if m:
            before = line[:m.start()].strip()
            after  = line[m.end():].strip()
            if before:
                fio_parts.append(before)
            if after:
                addr_parts.append(after)
            addr_started = True
        else:
            fio_parts.append(line)

    fio  = re.sub(r'\s+', ' ', ' '.join(fio_parts)).strip().rstrip(',.').strip()
    addr = '\n'.join(addr_parts).strip()

    # ── Пост-обработка ────────────────────────────────────────────────────────

    # 1) «ФИО СПб,...» — пробел перед «СПб» без запятой
    if not addr:
        m = re.search(r'\s+((?:спб\b|г\.\s).+)$', fio, re.IGNORECASE)
        if m:
            addr = m.group(1).strip()
            fio  = fio[:m.start()].strip()

    # 2) «ФИО. Улица 44-123» — точка перед адресом без ключевого слова города
    if not addr and re.search(r'\d{1,3}[-]\d{1,3}', fio):
        m = re.search(r'[,\.]\s+(\S.+\d{1,3}[-]\d{1,3}.*)$', fio)
        if m:
            addr = m.group(1).strip()
            fio  = fio[:m.start()].strip()

    # 3) «ФИО ул. Ушинского» — аббревиатура улицы оказалась в ФИО
    if re.search(r'\s+(?:ул\.|пр\.|пр-т)\s+\S', fio, re.IGNORECASE):
        m = re.search(r'\s+((?:ул\.|пр\.|пр-т)\s+.+)$', fio, re.IGNORECASE)
        if m:
            prefix = m.group(1).strip()
            addr = prefix + (' ' + addr if addr else '')
            fio  = fio[:m.start()].strip()

    # 4) В ФИО больше 3 слов и лишние слова — адресные компоненты
    #    («Алиев Сеймур Векилоглы Гражданский пр.» или «...пр» → trim)
    fio_words = fio.split()
    if len(fio_words) > 3:
        extra = ' '.join(fio_words[3:])
        if re.search(r'(?:пр|ул|д|кв|корп)\.?\b', extra, re.IGNORECASE):
            addr = extra + (' ' + addr if addr else '')
            fio  = ' '.join(fio_words[:3])

    return fio, addr


def clean_end_date(text: str) -> str:
    """Из ячейки «дд.мм.гггг\\nпримечание» вытаскивает только дату."""
    if not text:
        return ''
    first = text.split('\n')[0].strip()
    m = re.search(r'\d{1,2}[./]\d{1,2}[./]\d{2,4}', first)
    return m.group(0) if m else first


# ── Важные отметки в примечаниях ───────────────────────────────────────────────

_IMPORTANT_NOTE = re.compile(
    r'сво\b|сизо\b|умер(?:ла)?\b|розыск|скрылся|скрылась|'
    r'инвалид|побег|арест\b|задержан',
    re.IGNORECASE,
)


def extract_note(text: str) -> str:
    """Возвращает строки примечания, содержащие важные отметки; иначе ''."""
    if not text:
        return ''
    important = []
    for line in text.split('\n'):
        line = line.strip()
        if line and _IMPORTANT_NOTE.search(line):
            important.append(line)
    return '; '.join(important)


def get_b8_note(text: str) -> str:
    """Из ячейки «дата\\nпримечание» возвращает только часть после даты."""
    if not text:
        return ''
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    note_lines = [l for l in lines
                  if not re.search(r'^\d{1,2}[./]\d{1,2}[./]\d{2,4}', l)]
    return '\n'.join(note_lines)


# ── Нормализация полей ─────────────────────────────────────────────────────────

def normalize_date(text: str) -> str:
    """Нормализует дату к формату ДД.ММ.ГГГГ.

    Обрабатывает: пробел внутри («22.07. 25»), 2-значный год («04.03.25»),
    слэш-разделитель («03/07/2024»), суффиксы («04.03.25г.», «2025 года»).
    """
    if not text:
        return ''
    # Убираем суффиксы года
    text = re.sub(r'\s*(г\.?|года)\b', '', text.strip(), flags=re.IGNORECASE)
    # Убираем пробел внутри даты: «22.07. 25» → «22.07.25»
    text = re.sub(r'(\d{1,2}[./]\d{1,2}[./])\s+(\d{2,4})', r'\1\2', text)
    m = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{2,4})', text)
    if not m:
        return text.strip()
    day, month, year = m.group(1), m.group(2), m.group(3)
    if len(year) == 2:
        year = '20' + year
    return f'{int(day):02d}.{int(month):02d}.{year}'


def normalize_duties(text: str) -> str:
    """Объединяет строки обязанностей через «; »."""
    if not text:
        return ''
    # Снимаем ведущий дефис/тире, trailing «;», схлопываем внутренние пробелы
    lines = [re.sub(r'\s+', ' ', l.strip().lstrip('-–— ').rstrip(';')).strip() for l in text.split('\n')]
    lines = [l for l in lines if l]
    return '; '.join(lines)


# Паттерн суда: окончания прилагательных перед известными словами
_COURT_SPLIT = re.compile(
    r'(ским|ными?|овым|судом|ского|ному)(?=(район|суд|город|округ|ской|ского|г\.|р\.с|по\s))',
    re.IGNORECASE,
)

# Аббревиатуры суда (порядок важен: специфичные первыми)
_COURT_ABBR = [
    # «г. СПб» / «г. СПБ» → «г. Санкт-Петербурга» (до standalone-замены СПб)
    (re.compile(r'г\.?\s*СПб\b', re.IGNORECASE),   'г. Санкт-Петербурга'),
    # Типы судов (слэш и без)
    (re.compile(r'\bр/с\b',      re.IGNORECASE),   'районным судом'),
    (re.compile(r'\bрс\b',       re.IGNORECASE),   'районным судом'),   # без слэша
    (re.compile(r'\bг/с\b',      re.IGNORECASE),   'городским судом'),
    (re.compile(r'\bг\.с\.',     re.IGNORECASE),   'городским судом'),  # с точками
    (re.compile(r'\bв/с\b',      re.IGNORECASE),   'военным судом'),
    (re.compile(r'\bгар\.\s+',   re.IGNORECASE),   'гарнизонным '),     # военный
    # Город (standalone, уже без «г.»)
    (re.compile(r'\bСПб\b',      re.IGNORECASE),   'г. Санкт-Петербурга'),
    # Регионы
    (re.compile(r'\bЛО\b'),                        'Ленинградской области'),
    (re.compile(r'\bМО\b'),                        'Московской области'),
]


def normalize_court(text: str) -> str:
    """Разбивает склеенные слова в поле суда, раскрывает аббревиатуры."""
    if not text:
        return ''
    # 1. «2024г.» → «2024 г.»
    text = re.sub(r'(\d)(г\.)', r'\1 \2', text)
    # 2. Цифра + заглавная → добавить пробел  («2024Калининским», «ч.2УК»)
    text = re.sub(r'(\d)([А-ЯЁ])', r'\1 \2', text)
    # 3. «г.» + заглавная → добавить пробел  («г.Санкт»)
    text = re.sub(r'(г\.)([А-ЯЁ])', r'\1 \2', text)
    # 4. Строчная + заглавная (общий случай)
    text = _MERGE_CYR.sub(r'\1 \2', text)
    # 5. Окончания прилагательных суда + следующее известное слово
    text = _COURT_SPLIT.sub(r'\1 ', text)
    # 5б. Буква, слипшаяся с сокращением суда («Смольнтнскийр/с», «р/сг.»)
    text = re.sub(r'([а-яё])(р/с|г/с|в/с)', r'\1 \2', text, flags=re.IGNORECASE)
    text = re.sub(r'(р/с|г/с|в/с)([а-яё])', r'\1 \2', text, flags=re.IGNORECASE)
    # 5в. Убрать задвоенное «г. г.» (источник: «г . г. Санкт-Петербурга»)
    text = re.sub(r'\bг\s*\.\s+г\.', 'г.', text)
    # 6. Раскрываем аббревиатуры
    for pattern, replacement in _COURT_ABBR:
        text = pattern.sub(replacement, text)
    # 7. Прилагательное в именительном перед «районным/городским/etc. судом» → творительный
    #    (после раскрытия аббревиатур, чтобы «р/с» → «районным судом» уже было готово)
    #    «Калининский районным судом» → «Калининским районным судом»
    text = re.sub(
        r'\b([А-ЯЁа-яё]+)ий\s+((?:районным|городским|военным|областным)\s+судом)',
        r'\1им \2', text,
    )
    # 8. Прилагательное в родительном перед творительным судом → родительный суда
    #    «Калининского районным судом» → «Калининского районного суда»
    #    (бывает в формулировках «Постановлением X-ского районным судом»)
    text = re.sub(r'\b([А-ЯЁа-яё]+(?:ого|его))\s+районным\s+судом',  r'\1 районного суда',  text)
    text = re.sub(r'\b([А-ЯЁа-яё]+(?:ого|его))\s+городским\s+судом', r'\1 городского суда', text)
    text = re.sub(r'\b([А-ЯЁа-яё]+(?:ого|его))\s+военным\s+судом',   r'\1 военного суда',   text)
    return re.sub(r'\s+', ' ', text).strip()


# «г. СПб» / «г. СПБ» / «СПб» / «Спб» → «г. Санкт-Петербург»
_ADDR_SPB = re.compile(r'(?:г\.?\s*)?(?:СПб|СПБ|Спб)\b', re.IGNORECASE)

# Строки-телефоны: «Т.8-969-...», «(8-953-...)», «Т.: 8-994-...», «8-911-...»
_ADDR_PHONE = re.compile(
    r'^\(?'
    r'(?:[Тт](?:ел?)?\.?:?\s*)?'
    r'(?:\+7|8[-\s]?9)\d{2}[-\s\d]{5,}'
    r'[,\)]?\s*$'
)


def normalize_address_line(text: str) -> str:
    """Приводит адрес к одной строке: без переносов и лишних пробелов."""
    if not text:
        return ''
    return re.sub(r'\s+', ' ', str(text).replace('\r', ' ').replace('\n', ' ')).strip()


def normalize_address(text: str) -> str:
    """Нормализует адрес: убирает мусор (РФ, телефоны, гражданство, патроним, Инвалид)."""
    if not text:
        return ''
    # 1. Ведущий «РФ» (с запятой/точкой или без) до конца строки
    text = re.sub(r'^РФ[,\.]?\s*\n', '', text, flags=re.IGNORECASE)
    # 2. Фильтр строк — убираем «мусорные» строки целиком
    def _keep(line: str) -> bool:
        s = line.strip()
        if not s:
            return False
        if _ADDR_PHONE.match(s):                                    # телефон
            return False
        if re.match(r'^(?:гражданство\s+)?РФ[,.]?\s*$', s, re.IGNORECASE):  # «РФ», «Гражданство РФ»
            return False
        if re.match(r'^(?:гражданство\s+\w+|р\.\s+\w+|[Гг]р\.?\s+(?:[Рр]\.?\s+)?\w+)\s*$', s, re.IGNORECASE):  # «р. Узбекистан», «Гр. Р. Узбекистан»
            return False
        if re.match(r'^[Тт](?:ел?)?\.?:?\s*$', s):                # пустые «Тел.», «Т.:»
            return False
        if re.match(r'^[Ии]нвалид\b', s):                          # «Инвалид III гр.» — уже в Примечании
            return False
        return True
    lines = [l for l in text.split('\n') if _keep(l)]
    text = '\n'.join(lines)
    # 3. Инлайн: «, РФ» / «  РФ,» и всё последующее в строке (телефон и т.п.)
    text = re.sub(r',\s*РФ\b.*$', '', text, flags=re.IGNORECASE | re.MULTILINE)
    # 4. Инлайн: телефон с маркером «тел» / «т.:» («, тел89...», «. т.: 89...»)
    text = re.sub(r'[,.]?\s*[Тт](?:ел?)?\.?:?\s*(?:\+7|8[-\s]?9)[\d\s\-]{8,}',
                  '', text, flags=re.IGNORECASE)
    # 5. Инлайн: телефон в скобках «(89812123020)», «(8-953-171-14-51)»
    text = re.sub(r'\s*\((?:\+7|8)[-\s]?9[\d\s\-]{8,}\)', '', text)
    # 5б. Инлайн: телефон без маркера в конце строки («... 8-921-927-57-33»)
    text = re.sub(r'\s+8[-\s]?9\d{2}[-\s\d]{7,}$', '', text, flags=re.MULTILINE)
    # 6. СПб → г. Санкт-Петербург
    text = _ADDR_SPB.sub('г. Санкт-Петербург', text)
    # 7. Убрать патроним в начале («Толегенович г. Санкт-Петербург...»)
    text = re.sub(r'^[А-ЯЁ][а-яё]+(?:ович|евич|овна|евна)\s+', '', text)
    return normalize_address_line(text)


def normalize_record(rec: dict) -> dict:
    """Нормализует поля одной записи (изменяет на месте, возвращает rec)."""
    for fld in ('Дата постановки', 'Дата рождения', 'Окончание срока'):
        if rec.get(fld):
            rec[fld] = normalize_date(rec[fld])
    if rec.get('Обязанности'):
        rec['Обязанности'] = normalize_duties(rec['Обязанности'])
    if rec.get('Суд (когда, кем)'):
        rec['Суд (когда, кем)'] = normalize_court(rec['Суд (когда, кем)'])
    rec['Место жительства'] = normalize_address(rec.get('Место жительства', ''))
    return rec


def get_cell(row, idx: int) -> str:
    """Безопасно берёт текст ячейки (пустая строка если нет)."""
    try:
        return row.cells[idx].text.strip()
    except IndexError:
        return ''


def has_date(row) -> bool:
    """Признак строки с данными — есть хотя бы одна дата."""
    for cell in row.cells:
        if re.search(r'\d{1,2}[./]\d{1,2}[./]\d{2,4}', cell.text):
            return True
    return False


# ── Функции извлечения данных по типу структуры ────────────────────────────────

def _rec(category, date, num_ld, fio_raw, dob, addr_raw, court, duties, end, note=''):
    """Собирает словарь записи."""
    fio, addr_split = split_fio_address(fio_raw) if not addr_raw else (fio_raw, '')
    addr = addr_raw if addr_raw else addr_split
    return {
        'Категория':        category,
        'Дата постановки':  date,
        '№ л/д':            num_ld,
        'ФИО':              re.sub(r'\s+', ' ', fio).strip(),
        'Дата рождения':    dob,
        'Место жительства': normalize_address_line(addr),
        'Суд (когда, кем)': court,
        'Обязанности':      duties,
        'Окончание срока':  end,
        'Примечание':       note,
    }


def extract_A(table, data_start, category):
    """9 колонок: col1=дата, col2=№, col3=ФИО+адрес, col4=ДР,
                  col5=суд, col6=обязанности, col7=окончание, col8=примечание"""
    out = []
    for row in table.rows[data_start:]:
        if not has_date(row):
            continue
        out.append(_rec(
            category,
            date    = get_cell(row, 1),
            num_ld  = get_cell(row, 2),
            fio_raw = get_cell(row, 3),
            dob     = get_cell(row, 4),
            addr_raw= '',
            court   = get_cell(row, 5),
            duties  = get_cell(row, 6),
            end     = get_cell(row, 7),
            note    = extract_note(get_cell(row, 8)),
        ))
    return out


def extract_B8(table, data_start, category):
    """8 колонок: col7 = дата_окончания + примечание (объединено)."""
    out = []
    for row in table.rows[data_start:]:
        if not has_date(row):
            continue
        col7 = get_cell(row, 7)
        out.append(_rec(
            category,
            date    = get_cell(row, 1),
            num_ld  = get_cell(row, 2),
            fio_raw = get_cell(row, 3),
            dob     = get_cell(row, 4),
            addr_raw= '',
            court   = get_cell(row, 5),
            duties  = get_cell(row, 6),
            end     = clean_end_date(col7),
            note    = extract_note(get_b8_note(col7)),
        ))
    return out


def extract_no_end(table, data_start, category):
    """7-8 колонок без обязанностей и окончания (ОР, ИР).
    ИР (7 кол): col6 = примечание; ОР (8 кол): col6 может быть пустым."""
    out = []
    for row in table.rows[data_start:]:
        if not has_date(row):
            continue
        out.append(_rec(
            category,
            date    = get_cell(row, 1),
            num_ld  = get_cell(row, 2),
            fio_raw = get_cell(row, 3),
            dob     = get_cell(row, 4),
            addr_raw= '',
            court   = get_cell(row, 5),
            duties  = '',
            end     = '',
            note    = extract_note(get_cell(row, 6)),
        ))
    return out


def extract_otsr(table, data_start, category):
    """9 колонок (Отсрочка): col3=ФИО отдельно, col5=МЖ отдельно, col8=примечание."""
    out = []
    for row in table.rows[data_start:]:
        if not has_date(row):
            continue
        fio = re.sub(r'\s+', ' ', get_cell(row, 3)).strip()
        out.append(_rec(
            category,
            date    = get_cell(row, 1),
            num_ld  = get_cell(row, 2),
            fio_raw = fio,         # уже чистый ФИО
            dob     = get_cell(row, 4),
            addr_raw= get_cell(row, 5),   # отдельная колонка
            court   = get_cell(row, 6),
            duties  = '',
            end     = get_cell(row, 7),
            note    = extract_note(get_cell(row, 8)),
        ))
    return out


def extract_da(table, data_start, category):
    """8 колонок (Домашний арест): col6=обязанности, col7=примечание."""
    out = []
    for row in table.rows[data_start:]:
        if not has_date(row):
            continue
        out.append(_rec(
            category,
            date    = get_cell(row, 1),
            num_ld  = get_cell(row, 2),
            fio_raw = get_cell(row, 3),
            dob     = get_cell(row, 4),
            addr_raw= '',
            court   = get_cell(row, 5),
            duties  = get_cell(row, 6),
            end     = '',
            note    = extract_note(get_cell(row, 7)),
        ))
    return out


def extract_zoda(table, data_start, category):
    """8 колонок (ЗОДА): col5 = суд + обязанности вместе, col7=примечание."""
    out = []
    for row in table.rows[data_start:]:
        if not has_date(row):
            continue
        out.append(_rec(
            category,
            date    = get_cell(row, 1),
            num_ld  = get_cell(row, 2),
            fio_raw = get_cell(row, 3),
            dob     = get_cell(row, 4),
            addr_raw= '',
            court   = get_cell(row, 5),   # суд + запреты вместе
            duties  = '(см. суд)',
            end     = '',
            note    = extract_note(get_cell(row, 7)),
        ))
    return out


# ── Разбор документа ───────────────────────────────────────────────────────────

def parse_document(doc_path: str):
    doc = Document(doc_path)

    # Проходим по телу документа и устанавливаем категорию для каждой таблицы
    SKIP_TEXTS = {'С П И С О К', 'инспектор', 'Телефон', 'Подъячева'}

    category = 'Условное осуждение'
    table_cats = {}
    t_count = 0

    for elem in doc.element.body:
        tag = elem.tag.split('}')[-1]
        if tag == 'p':
            text = ''.join(n.text or '' for n in elem.iter(qn('w:t'))).strip()
            # Обновляем категорию только по коротким «заголовочным» параграфам
            if text and t_count > 0 and not any(s in text for s in SKIP_TEXTS):
                category = text
        elif tag == 'tbl':
            table_cats[t_count] = category
            t_count += 1

    # Первая таблица — до любых параграфов-категорий
    table_cats[0] = 'Условное осуждение'

    # Маппинг индекса таблицы → функция извлечения + начало данных
    TABLE_HANDLERS = {
        0:  (extract_A,      1),   # УС: заголовок в строке 0
        1:  (extract_A,      0),   # Принудительное лечение
        2:  (extract_A,      0),   # Штраф с лечением
        3:  (extract_B8,     0),   # Несовершеннолетние
        4:  (extract_A,      0),   # ЗЗД
        5:  (extract_no_end, 0),   # Обязательные работы
        6:  (extract_no_end, 1),   # Исправительные работы (строка 0 — merged-заголовок)
        7:  (extract_no_end, 0),   # Исправительные работы (продолжение)
        8:  (extract_otsr,   1),   # Отсрочка (строка 0 — заголовок, ФИО/МЖ раздельно)
        9:  (extract_da,     0),   # Домашний арест
        10: (extract_zoda,   0),   # ЗОДА
        11: (extract_B8,     0),   # Ограничение свободы
        12: (extract_B8,     0),   # УДО
    }

    all_records = []
    for t_idx, table in enumerate(doc.tables):
        if t_idx not in TABLE_HANDLERS:
            print(f'  [!] Таблица {t_idx} не имеет обработчика — пропущена')
            continue
        handler, start = TABLE_HANDLERS[t_idx]
        cat = table_cats.get(t_idx, 'Неизвестно')
        records = handler(table, start, cat)
        print(f'  Таблица {t_idx:2d} | {cat[:40]:40s} | {len(records)} записей')
        all_records.extend(records)

    for rec in all_records:
        normalize_record(rec)

    return all_records


# ── Запись в Excel ─────────────────────────────────────────────────────────────

COLUMNS = [
    'Категория',
    'Дата постановки',

    'ФИО',
    'Дата рождения',
    'Место жительства',
    'Суд (когда, кем)',
    'Обязанности',
    'Окончание срока',
    'Примечание',
]

# Пастельные цвета для каждой категории
CATEGORY_COLORS = {
    'Условное осуждение':                                  'BBDEFB',
    'ПРИНУДИТЕЛЬНОЕ лечение':                              'E1BEE7',
    'ШТРАФ С ЛЕЧЕНИЕМ (СТ. 72.1 ук РФ)':                 'FFE0B2',
    'несовершеннолетние':                                  'C8E6C9',
    'ЗЗД':                                                 'FFF9C4',
    'ОБЯЗАТЕЛЬНЫЕ РАБОТЫ':                                 'FFCCBC',
    'Отсрочка до достижения ребенком 14-летнего возраста': 'DCEDC8',
    'Домашний арест ЕЖЕМЕСЯЧНО':                           'B2EBF2',
    'ЗОДЕЖЕМЕСЯЧНО':                                       'D1C4E9',
    'Ограничение свободы':                                 'FFCDD2',
    'УДО':                                                 'C5CAE9',
}


def write_excel(records: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Список УИИ'

    # ─ Заголовок ──────────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_i, col_name in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_i, value=col_name)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = hdr_align

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}1'

    # ─ Данные ─────────────────────────────────────────────────────────────────
    thin   = Side(style='thin', color='BDBDBD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_align = Alignment(vertical='top', wrap_text=True)

    current_cat   = None
    current_color = 'FFFFFF'

    for row_i, rec in enumerate(records, 2):
        cat = rec.get('Категория', '')
        if cat != current_cat:
            current_cat   = cat
            current_color = CATEGORY_COLORS.get(cat, 'F5F5F5')

        row_fill = PatternFill(start_color=current_color, end_color=current_color,
                               fill_type='solid')

        for col_i, col_name in enumerate(COLUMNS, 1):
            c = ws.cell(row=row_i, column=col_i, value=rec.get(col_name, ''))
            c.alignment = data_align
            c.fill      = row_fill
            c.border    = border

    # ─ Ширина колонок ─────────────────────────────────────────────────────────
    widths = [24, 14, 11, 28, 12, 32, 48, 42, 14]
    for col_i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    wb.save(output_path)
    print(f'\nГотово! Файл сохранён: {output_path}')
    print(f'Всего записей: {len(records)}')


# ── Запуск ─────────────────────────────────────────────────────────────────────

def main():
    sys.stdout.reconfigure(encoding='utf-8')

    if not os.path.exists(INPUT_FILE):
        print(f'Ошибка: файл {INPUT_FILE!r} не найден.')
        print(f'Запустите скрипт из папки с docx-файлом.')
        sys.exit(1)

    print(f'Читаем: {INPUT_FILE}')
    print()
    records = parse_document(INPUT_FILE)
    print()
    write_excel(records, OUTPUT_FILE)


if __name__ == '__main__':
    main()
